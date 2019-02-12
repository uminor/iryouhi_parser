#*********************************************
# S社　健康保険医療費 集計プログラム Ver.1.01
#*********************************************

# ≪開発者≫ +++++++++++++++++++++++++++++++++++++++++++++++++++++++
# Nuts Westfield (uminor)
# 参考にさせていただいたソース：
#  https://qiita.com/mczkzk/items/894110558fb890c930b5
#
# ≪概要≫ +++++++++++++++++++++++++++++++++++++++++++++++++++++++++
# 1年分の医療費の自己負担額が10万円を超える場合、
# 確定申告すれば医療費控除が可能。
#
# しかしS社健保組合が発行している「健康保険医療費のお知らせ」
#　の pdfファイル（以下pdf）は、
#
# （１）紙に印刷しただけの状態では「原本」と言えないので、
# 　　医療費控除の正式な証票となるかどうか疑問。
#     かといって pdf そのものはパスワードかかかっており、
#     e-taxの申告時に添付するのは不適切。
#
# （２）仮に証票として使えるとしても、e-tax で申告した場合
#    「別途郵送」の手間が発生する。
#
# そこで、本プログラムでは、
#
#　(a)各領収書は別途保管すること。
#  (b)pdf の内容は各領収書と一致していること。
#
# を前提として、pdf を「集計」のための元データとして用いる。
#
# 本プログラムは、pdf 約1年分を読み込んで、
# 集計したExcelファイルを出力する。
# 但し、そのままでは、e-tax システムで読めないので、
# それを開いて、医療費控除申告用の「医療費集計フォーム」
# (国税庁ホームページで配布)に手動でコピペする。
# この「医療費集計フォーム」を e-tax システムで読み込ば、、
# 申告で用いる「医療費控除の明細書」に反映する。
# 
# ≪注意事項≫ +++++++++++++++++++++++++++++++++++++++++++++++++++++
# 本プログラムの動作は無保証です。
# 特に、入力データに依存して正しく動作しない可能性があり、
# 正しい申告データが作成されるかどうかも無保証です。
# （動作レポート・バグ報告は歓迎します。）
#
# ≪使い方≫ +++++++++++++++++++++++++++++++++++++++++++++++++++++++
# １．下記「必要な環境」をインストールする。
#
# ２．本プログラム(iryohi_parser.py)と同じフォルダに
# 　サブフォルダを作り、S社サイトから取得した
# 「健康保険医療費のお知らせ」の pdf 1年分を入れる。
#  (通常、少なくとも12月分は間に合わない ;_;)
#
# ３．コマンドプロンプトを起動し
#  cd [iryouhi_parser.py があるパス]
#  python iryouhi_parser.py [フォルダ名] [pdfパスワード]
#  （例： python iryouhi_parser.py pdfs xxxyyyzzz）
#　で、本プログラムを起動する。
#
# ４．処理が正常に完了すると、Excelファイル：
# 「これを開いて医療費集計フォームにコピペ.xlsx」が出力される。
#
# ５． 但し、そのままでは、e-tax システムで読めないので、
# 　それを開いて、医療費控除申告用の「医療費集計フォーム」
#　　（iryouhi_form_v3.xlsx　＝国税庁ホームページで配布)
#    の該当部分にコピペする。
#
# ６．この「医療費集計フォーム」を e-tax システムで読み込ば、、
# 　申告で用いる「医療費控除の明細書」に反映する。
#
# ≪必要な環境≫ +++++++++++++++++++++++++++++++++++++++++++++++++++
#
# (1)Windows7 64bit 　(で動作確認)
#
# (2)Python3（コンピュータ言語）, pip3（ライブラリ管理システム）
#  インストール方法：略
#
# (3)ライブラリ pdfminer(pdf読み込み), openpyxl(xlsx読み書き)
#  インストール方法：
#    pip3 install pdfminer.six
#    pip3 install openpyxl
#
# (4)qpdf コマンド（pdfパスワード解除）
#  インストール方法：
#  https://sourceforge.net/projects/qpdf/files/qpdf/
# から 8.4.0/qpdf-8.4.0-bin-msvc64.zip を取得して解凍し、
# qpdf.exe　のパスで次の行の右辺を書き換える。

qpdf_path = 'C:\\Users\\minor\\Documents\\x\\qpdf-6.0.0\\bin\\qpdf.exe'

wk_pdf = "_wk_.pdf"
vorbose = False
import sys
import os
import subprocess
import inspect
import openpyxl
from pdfminer.converter import PDFPageAggregator
from pdfminer.layout import LAParams, LTContainer, LTTextBox
from pdfminer.pdfinterp import PDFPageInterpreter, PDFResourceManager
from pdfminer.pdfpage import PDFPage

qpdf_path = 'C:\\Users\\minor\\Documents\\x\\qpdf-6.0.0\\bin\\qpdf.exe'

# 参考ソース： https://qiita.com/mczkzk/items/894110558fb890c930b5
def find_textboxes_recursively(layout_obj):
    """
    再帰的にテキストボックス（LTTextBox）を探して、テキストボックスのリストを取得する。
    """
    # LTTextBoxを継承するオブジェクトの場合は1要素のリストを返す。
    if isinstance(layout_obj, LTTextBox):
        return [layout_obj]

    # LTContainerを継承するオブジェクトは子要素を含むので、再帰的に探す。
    if isinstance(layout_obj, LTContainer):
        boxes = []
        for child in layout_obj:
            boxes.extend(find_textboxes_recursively(child))

        return boxes

    return []  # その他の場合は空リストを返す。

# Layout Analysisのパラメーターを設定。縦書きの検出を有効にする。
laparams = LAParams(detect_vertical=True)

# 共有のリソースを管理するリソースマネージャーを作成。
resource_manager = PDFResourceManager()

# ページを集めるPageAggregatorオブジェクトを作成。
device = PDFPageAggregator(resource_manager, laparams=laparams)

# Interpreterオブジェクトを作成。
interpreter = PDFPageInterpreter(resource_manager, device)

# 出力用のテキストファイル
output_txt = open('output.txt', 'w')

def print_and_write(txt):
    print(txt)
    output_txt.write(txt)
    output_txt.write('\n')

def include(ky, lis):
    for l in lis:
        if ky in l:
            return True
    return False

def parse_pdf(file, total_):
    with open(file, 'rb') as f:
        # PDFPage.get_pages()にファイルオブジェクトを指定して、PDFPageオブジェクトを順に取得する。
        # 時間がかかるファイルは、キーワード引数pagenosで処理するページ番号（0始まり）のリストを指定するとよい。
        page_num = 0
        for page in PDFPage.get_pages(f):
            page_num += 1
            if vorbose:
                print_and_write('\n====== ページ {0} ======\n'.format(page_num))
            interpreter.process_page(page)  # ページを処理する。
            layout = device.get_result()  # LTPageオブジェクトを取得。

            # ページ内のテキストボックスのリストを取得する。
            boxes = find_textboxes_recursively(layout)

            # テキストボックスの左上の座標の順でテキストボックスをソートする。
            # y1（Y座標の値）は上に行くほど大きくなるので、正負を反転させている。
            ##boxes.sort(key=lambda b: (-b.y1, b.x0))

            for box in boxes:
                vkey = int(round((box.y1)/38,0))
                pvkey = page_num * 100 + vkey
                hkey = int(box.x0) # 153 --> 141
                content = box.get_text().strip().replace('\n','')  # テキストボックス内のテキストを表示する。
                if vorbose:
                    print_and_write(content + "({0},{1}):{2}".format(round(box.y1,0), round(box.x0,0), vkey))
                if 4 <= vkey <= 13:
                    if not (pvkey in dic):
                        dic[pvkey] = []
                    if 1 <= len(content) < 64 and not include("計", dic[pvkey]):
                        if 153 == hkey and 3 == len(dic[pvkey]):
                            dic[pvkey][2] += content # 病院名が２行に分かれている場合
                        else:
                            dic[pvkey].append(content)
                            if 10 == len(dic[pvkey]):
                                total_ += int(dic[pvkey][9].replace(',',''))
    return total_

dir = sys.argv[1]
pw = sys.argv[2]

files = os.listdir(dir)


def fmt_rec(ar):
    return '"' + '","'.join(ar) + '"'

class Meisai:
    def __init__(self, one):
        self.patient = one[0].replace("　","")
        self.month   = one[1]
        self.days, self.hospital = one[2].split()
        self.kind    = one[3]
        self.sogaku  = int(one[4].replace(',', ''))
        self.kenpo_futan = int(one[5].replace(',', ''))
        self.kohi_futan  = int(one[6].replace(',', ''))
        self.jiko_futan  = int(one[7].replace(',', ''))
        self.kenpo_kyufu = int(one[8].replace(',', ''))
        self.real_futan  = int(one[9].replace(',', ''))
    
    def put_sheet(self, sht, num):
        ro = 8 + num
        sht.cell(row=ro, column=2, value=self.patient)
        sht.cell(row=ro, column=3, value=self.hospital)
        if '通院' == self.kind:
            sht.cell(row=ro, column=4, value='該当する')
        if '薬局' == self.kind:
            sht.cell(row=ro, column=5, value='該当する')
        sht.cell(row=ro, column=8, value=self.jiko_futan)
        sht.cell(row=ro, column=9, value=self.kenpo_kyufu)
        sht.cell(row=ro, column=10, value=self.month)

c = 0
total = 0
rows = 0
meisais = []
for i, f in enumerate(files):
    if os.path.exists(wk_pdf):
        os.remove(wk_pdf)
    dic = {}
    print('----------------\nfile {0} {1}'.format(i,f))
    cmd = "{0} --decrypt {1} --password={2} {3}".format(qpdf_path, dir + '/' + f, pw, wk_pdf)
    print(cmd)
    returncode = subprocess.call(cmd)
    total = parse_pdf(wk_pdf, total)
    print()
    c = i

    for d in dic:
        if not include('計', dic[d]) and 10 == len(dic[d]):
            print_and_write("({0}){1}".format(d, fmt_rec(dic[d])))
            meisais.append(Meisai(dic[d]))
            rows += 1
            #print(inspect.getmembers(meisai))

print('\nParsed {0} files. Total={1}'.format(c + 1, total))

wb = openpyxl.load_workbook('医療費集計テンプレート.xlsx')
sheet = wb['Sheet1']

#wb = openpyxl.load_workbook('iryouhi_form_v3_resave.xlsx')
#wb = openpyxl.load_workbook('iryouhi_form_v3_sanitized.xlsx')
#sheet = wb['医療費集計フォーム']

for i, meisai in enumerate(meisais):
    meisai.put_sheet(sheet, i+1)

wb.save('これを開いて医療費集計フォームにコピペ.xlsx')

output_txt.close()

